# Imports
import openpyxl
import pandas as pd
import numpy as np
from sklearn.compose import ColumnTransformer
from sklearn.preprocessing import OneHotEncoder, StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestRegressor, BaggingRegressor
from sklearn.svm import SVR
from sklearn.metrics import r2_score
import matplotlib.pyplot as plt

# Helper functions
def get_column(sheet_obj, start_raw, column):
    raw_num = start_raw
    lst = []
    
    while True:
        cell_obj = sheet_obj.cell(row=raw_num, column=column)
        if cell_obj.value is None:
            break
        lst.append(cell_obj.value)
        raw_num += 1
    
    return lst

def get_raw(sheet_obj, start_column, raw, num_columns):
    col_num = start_column
    num_columns += start_column
    lst = []
    
    while True:
        cell_obj = sheet_obj.cell(row=raw, column=col_num)
        if num_columns == col_num:
            break
        lst.append(cell_obj.value)
        col_num += 1
    
    return lst

# Read rainfall and temperature data
start_raw = 5
start_column = 2
num_columns = 17
price_excel_path = "rain_temp_fuel_2021.xlsx"

wb_obj = openpyxl.load_workbook(price_excel_path)
sheet_name = wb_obj.sheetnames[1]
sheet_obj = wb_obj[sheet_name]

title_lst = get_column(sheet_obj, start_raw, start_column)

temp_data = dict()
rain_data = dict()
district = ''
districts = []

for i, item in enumerate(title_lst):
    if i % 14 == 0:
        # Convert to string first and handle None values
        if item is not None:
            district = str(item).strip().lower()
            districts.append(district)
            rain_data[district] = []
            temp_data[district] = []
    elif (i-1) % 14 != 0 and district:  # Only proceed if we have a valid district
        rain_data[district].append(get_raw(sheet_obj, start_column+2, i+start_raw, 7))
        temp_data[district].append(get_raw(sheet_obj, start_column+10, i+start_raw, 7))

# Rest of the code remains the same...

# Process missing values in rainfall and temperature data
for district in districts:
    for i, lst in enumerate(rain_data[district]):
        avg = 0
        num_items = 0
        for j, item in enumerate(lst):
            val = 0
            try:
                val = float(item)
                num_items += 1
            except:
                pass
            avg += val
        avg /= num_items
        for j, item in enumerate(lst):
            try:
                float(item)
            except:
                rain_data[district][i][j] = round(avg, 2)

for district in districts:
    for i, lst in enumerate(temp_data[district]):
        avg = 0
        num_items = 0
        for j, item in enumerate(lst):
            val = 0
            try:
                val = float(item)
                num_items += 1
            except:
                pass
            avg += val
        avg /= num_items
        for j, item in enumerate(lst):
            try:
                float(item)
            except:
                temp_data[district][i][j] = round(avg, 2)

# Read crop price data
start_raw = 5
year_column = 2
num_columns = 15
price_excel_path = "Wholesale Prices of Selected Food Crops 2015-2021.xlsx"

wb_obj = openpyxl.load_workbook(price_excel_path)
sheet_name = wb_obj.sheetnames[0]
sheet_obj = wb_obj[sheet_name]

table_col_names = get_raw(sheet_obj, year_column, 4, num_columns)
years = get_column(sheet_obj, start_raw, year_column)

data = []
for i in range(start_raw, len(years) + start_raw):
    data.append(get_raw(sheet_obj, year_column, i, num_columns))

# Clean price data
data_new = []
for i in data:
    if i[0] < 2015 and i[0] > 2021:
        continue
    if i[2] is None:
        continue
    
    i[1] = i[1].lower()
    i[2] = i[2].lower()
    
    appendable = True
    for j in range(3, num_columns):
        p = 0
        try:
            p = float(i[j])
        except:
            appendable = False
            break
        if p < 30:
            appendable = False
            break
    
    if appendable:
        data_new.append(i)

# Process items
items = []
item_vals = []
for i in data_new:
    if i[2] not in items:
        items.append(i[2])
        item_vals.append([])

items_dict = dict(zip(items, item_vals))

for i in data_new:
    items_dict[i[2]].append(i)

# Read fuel data
price_excel_path = "Fuel_Prices.xlsx"
sheet_name = 'fuel'
months = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
fuel_types = ["LP95", "LP92", "LAD", "LSD", "LK", "LIK"]

wb_obj = openpyxl.load_workbook(price_excel_path)
sheet_obj = wb_obj[sheet_name]

fuel_data = dict()
column = 0
flag = True
year = 2014
month_i = 0

while flag:
    month = months[month_i]
    
    for i, ft in enumerate(fuel_types):
        cell_obj = sheet_obj.cell(row=i+1, column=column+1)
        
        if cell_obj.value is None:
            flag = False
            break
        
        fuel_type = fuel_types[i]
        
        if fuel_type not in fuel_data:
            fuel_data[fuel_type] = dict()
        if year not in fuel_data[fuel_type]:
            fuel_data[fuel_type][year] = dict()
            
        fuel_data[fuel_type][year][month] = float(cell_obj.value)
    
    month_i += 1
    if month_i == 12:
        month_i = 0
        year += 1
    
    column += 1

# Create main data table
main_table = pd.DataFrame(columns=['year', 'item', 'district', 'month', 'temperature', 'rain',
                                 'LP95', 'LP92', 'LAD', 'LSD', 'LK', 'LIK', 'price'])

for item_name in items_dict:
    for lst in items_dict[item_name]:
        year = lst[0]
        district = lst[1]
        for i in range(3, 15):
            district = district.strip().lower()
            month = table_col_names[i].strip().lower()
            year = int(year)
            lp95 = fuel_data['LP95'][year][month]
            lp92 = fuel_data['LP92'][year][month]
            lad = fuel_data['LAD'][year][month]
            lsd = fuel_data['LSD'][year][month]
            lk = fuel_data['LK'][year][month]
            lik = fuel_data['LIK'][year][month]
            
            if district in districts:
                temperature = temp_data[district][months.index(month)][year-2015]
                rain = rain_data[district][months.index(month)][year-2015]
                df_raw = {'year': year, 'item': item_name.strip().lower(), 'district': district, 'month': month,
                         'temperature': temperature, 'rain': rain, 'price': lst[i],
                         'LP95': lp95, 'LP92': lp92, 'LAD': lad, 'LSD': lsd, 'LK': lk, 'LIK': lik}
                main_table = main_table.append(df_raw, ignore_index=True)

# Remove 2021 data for training
main_table_for_train = main_table[main_table['year'] != 2021]

# Prepare data for modeling
x = main_table_for_train.iloc[:, 1:-1].values
y = main_table_for_train.iloc[:, -1].values

# Feature engineering
ct = ColumnTransformer(transformers=[('encoder', OneHotEncoder(), [0, 1, 2])], remainder='passthrough')
x = ct.fit_transform(x)
x = x.toarray()

# Scale features
sc_x = StandardScaler()
sc_y = StandardScaler()
sx = sc_x.fit_transform(x)
sy = sc_y.fit_transform(y.reshape(-1, 1))

# Split data
x_train, x_test, y_train, y_test = train_test_split(x, y, test_size=0.1, random_state=42)
x_sc_train, x_sc_test, y_sc_train, y_sc_test = train_test_split(sx, sy, test_size=0.1, random_state=42)

# Train models
regressor = RandomForestRegressor(n_estimators=25, random_state=42)
regressor.fit(x_train, y_train)

bagging_regressor = BaggingRegressor(base_estimator=SVR(), n_estimators=25, random_state=42)
bagging_regressor.fit(x_sc_train, y_sc_train.reshape(-1))

# Model evaluation
y_pred = regressor.predict(x_test)
bagging_y_pred = bagging_regressor.predict(x_sc_test)

rf_r2 = r2_score(y_test, y_pred)
bagging_r2 = r2_score(y_test, sc_y.inverse_transform(bagging_y_pred.reshape(-1, 1)))

print(f"Random Forest R2 Score: {rf_r2}")
print(f"Bagging with SVR R2 Score: {bagging_r2}")

# Example prediction
def make_prediction(crop, district, month, rain_temp_fuel_data):
    data_input = [[crop, district, month,
                   rain_temp_fuel_data[district]['temp'][month],
                   rain_temp_fuel_data[district]['rain'][month],
                   rain_temp_fuel_data[district]['LP95'][month],
                   rain_temp_fuel_data[district]['LP92'][month],
                   rain_temp_fuel_data[district]['LAD'][month],
                   rain_temp_fuel_data[district]['LSD'][month],
                   rain_temp_fuel_data[district]['LK'][month],
                   rain_temp_fuel_data[district]['LIK'][month]]]
    
    rf_pred = regressor.predict(ct.transform(data_input).toarray())
    bagging_pred = sc_y.inverse_transform(
        bagging_regressor.predict(
            sc_x.transform(ct.transform(data_input).toarray())
        ).reshape(-1, 1)
    )
    
    return rf_pred[0], bagging_pred[0][0]

# Example usage:
# rf_prediction, bagging_prediction = make_prediction('samba 1', 'colombo', 'jan', new_rain_temp_fuel)